import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
from get_setting import get_setting
from util import read_datafile
from tel_send import tel_send

inv_days = get_setting('inv_days', default='')  # 투자일수 
vo_rate = get_setting('vo_rate', default='')    # vo_rate = 0.5   # 변동성 돌파 적용 비율

# 엑셀 파일 셀에 시뮬레이션 결과 자료 타이틀 입력
def set_title(ws, base_cell, code, name, days, rate):
    title = ['No.','날짜','시가','고가','저가','현재가','5MA','10MA','매수가','당일수익','익일수익','수익률','MA미적용']

    ws['A1'] = '종목코드'    
    ws['B1'] = code
    ws['C1'] = '종목명'
    ws['D1'] = name
    ws['D1'].font = Font(name='Arial', size=12, bold=True, color='FF0000')  # 폰트
    ws['A2'] = '적용일수'
    ws['B2'] = days
    ws['C2'] = '변동률'
    ws['D2'] = rate
    ws['E2'] = '매수일수'
    ws['F2'] = 0
    ws['G2'] = '평균매수가'
    ws['H2'] = 0
    ws['I2'] = '평균수익률'
    ws['J2'] = 0.00
    ws['J2'].font = Font(name='Arial', size=12, bold=True, color='FF0000')  # 폰트
    

    # 기준 셀 설정
    act_cell = base_cell    

    for i in range(0, len(title)):
        act_cell = base_cell.offset(row=0, column=i)    
        act_cell.value = title[i]
        act_cell.font = Font(bold=True) 



# N일 이동평균 구하기
def moving_average(start, days, rows):
    ma_sum = 0
    for i in range(0, days):
        if (i < days):
            ma_sum += int(rows[start]['cur_prc']) 
            start += 1   
    ma = int(ma_sum/days)
    # print(f'{days}일 이동평균: {ma:,}원')
    return ma


    
    # 3. 새로운 시트 생성
    new_ws = wb.create_sheet(sheet_name)


def cal_roi(days, ws=None, rows=None, wb=None):
    vo_lastday, o_p, l_p, h_p, c_p = 0, 0, 0, 0, 0
    roi_sum, buy_sum, today_roi_sum, buy_count = 0, 0, 0, 0

    base_cell = ws['A4']

    for j in range(1, days+1):
        # 전일 변동성 = 전일 최고가 - 전일 최저가
        vo_lastday = int(rows[j]['high_pric']) - int(rows[j]['low_pric'])           
        # 당일 시가 + (전일 고가 - 전일 저가) * rate(0.5)
        target_price = int(rows[j-1]['open_pric']) + int(vo_lastday * vo_rate)

        o_p = int(rows[j-1]['open_pric'])   # 당일 시가
        l_p = int(rows[j-1]['low_pric'])    # 당일 최저가
        h_p = int(rows[j-1]['high_pric'])   # 당일 최고가    
        c_p = int(rows[j-1]['cur_prc'])     # 당일 현재가(종가)
        print(f'일자 {int(rows[j-1]['dt'])} 시가 {o_p:,} 최고 {h_p:,} 저가 {l_p:,} 종가 {c_p:,} 매수 {target_price:,}', end=' ')

        ma5 = moving_average(j-1, 5, rows)
        ma10 = moving_average(j-1, 10, rows)
        
        # 엑셀 cell 값 입력 시작        
        act_cell = base_cell
        act_cell.value = j                              # 번호 
        act_cell = base_cell.offset(row=0, column=1)
        act_cell.value = rows[j-1]['dt']                # 일자
        act_cell = base_cell.offset(row=0, column=2)
        act_cell.value = o_p                            # 시가
        act_cell = base_cell.offset(row=0, column=3)
        act_cell.value = h_p                            # 최고가
        act_cell = base_cell.offset(row=0, column=4)
        act_cell.value = l_p                            # 최저가
        act_cell = base_cell.offset(row=0, column=5)
        act_cell.value = c_p                            # 종가(현재가)
        act_cell = base_cell.offset(row=0, column=6)
        act_cell.value = ma5                            # MA5
        act_cell = base_cell.offset(row=0, column=7)
        act_cell.value = ma10                           # MA10
        act_cell = base_cell.offset(row=0, column=8)
        act_cell.value = target_price                   # 매수가
        # 엑셀 cell 값 입력 끝


        # 당일 시가가 매수가보다 높고(변동성 돌파), 5일 이평 및 10일 이평 이상일 때 매수
        if h_p > target_price and ma5 < target_price and ma10 < target_price:
            today_roi = c_p - target_price      # 당일 종가 - 매수가
            buy_sum += target_price             # 총 매수금액
            buy_count += 1                      # 총 매입일수
            today_roi_sum += today_roi          # 총 당일수익금액
                    

            print(f'당일익: {today_roi}')
            # 엑셀 cell 입력
            act_cell = base_cell.offset(row=0, column=9)    # 당일수익
            act_cell.value = today_roi
            act_cell = base_cell.offset(row=0, column=10)   # 익일수익
            act_cell.value = today_roi              
            if j > 1 and today_roi > 0:                     # 수익 중이면 익일 시가로 매도
                today_roi = int(rows[j-2]['open_pric']) - target_price
                print(f'익일익: {today_roi}')
                # 엑셀 cell 입력
                act_cell = base_cell.offset(row=0, column=10)   # 익일수익
                act_cell.value = today_roi  
                roi_sum += today_roi                        # 총 수익금액
            else:
                roi_sum += today_roi                        # 총 수익금액
        else:
            print('매수불가')
            # 엑셀 cell 입력
            act_cell = base_cell.offset(row=0, column=9)
            act_cell.value = False 
            act_cell = base_cell.offset(row=0, column=10)
            act_cell.value = False 
            today_roi = 0
        
        act_cell = base_cell.offset(row=0, column=11)       # 수익률
        act_cell.value =  round((today_roi / target_price) * 100, 2)

        # 엑셀 cell 다음 행 첫번째 셀로
        base_cell = ws[f'A{4+j}']

    if buy_count:
        abc = int(buy_sum / buy_count)               # 평균 매수가
        arr = round((roi_sum/ abc) * 100, 2)         # 평균 수익률
    else:
        abc, arr = 0, 0

    print('---------------------------------------------------------------------------------------')
    print(f'매수일수: {buy_count}일 평균수익률: {arr}% 평균매수가: {abc:,}원 수익금액: {roi_sum:,}원')
    
    # 엑셀 파일 저장
    ws['F2'] = buy_count                            # 총 매수일수
    ws['H2'] = abc                                  # 평균 매수가 average buying cost
    act_cell = base_cell.offset(row=0, column=0)
    act_cell = ws['J2']
    act_cell.value = arr                            # 평균 수익률  average return rate
    act_cell = base_cell.offset(row=0, column=0)
    act_cell.value = '합  계'
    act_cell = base_cell.offset(row=0, column=8)
    act_cell.value = buy_sum                        # 총 매수금액
    act_cell = base_cell.offset(row=0, column=9)    # 총 당일수익
    act_cell.value = today_roi_sum
    act_cell = base_cell.offset(row=0, column=10)   # 총 익일수익
    act_cell.value = roi_sum              

    sum_list = [abc, arr, roi_sum]

    # 엑셀 파일 저장
    wb.save('시뮬레이션결과.xlsx')

    return sum_list # roi_sum

async def vobres_sim_excel(filename, option):
    total_sum = []
    abc, arr, roi = 0,0,0

    script_dir = os.path.dirname(os.path.abspath(__file__))
    f_path = os.path.join(script_dir, filename)

    # excel_fname = filename#"주식종목및코드 260510.xlsx"
    # # sh_name = "코스닥20"
    # sh_name = option#"코스피100"

    df = pd.read_excel(f_path, sheet_name=option, engine='openpyxl')

    names = df['종목']      # 종목열만 추출 
    #print(type(names))    # names의 type() <class 'pandas.Series'>

    for i in range(0,1):#len(names)):
        code_n_name = names[i]
        code = code_n_name[-6:]     # 종목코드 추출
        name = code_n_name[:-6]     # 종목명 추출
        print(f'\n{i+1} 종목코드: {code} 종목명: {name}')
    
        # 주식번호로 일봉차트 데이터 파일 읽기
        response = read_datafile(name+'.txt', 'daypole_data') # response type <class 'str'> 내용은 dict(json) 문자열
        # json_data type <class 'dict'>
        json_data = json.loads(response)    # 파이썬 객체 str response를 dict json_data로 변환 (파이썬에서 json type은 따로 없으므로 dict 객체로 변환)         

        rows = json_data.get('stk_dt_pole_chart_qry', [])   # rows = json_data['stk_dt_pole_chart_qry']	동일함
        #print(f'테이터 type()  response: {type(response)}, json_data: {type(json_data)}, rows: {type(rows)}')

        # 1. 엑셀 파일 불러오기
        try:
            wb = openpyxl.load_workbook("시뮬레이션결과.xlsx")
        except FileNotFoundError:       
            wb = openpyxl.Workbook()        # 파일이 없으면 새로 생성
            print("시뮬레이션결과.xlsx 파일이 없으면 새로 생성")            
            if 'Sheet' in wb.sheetnames:    # 기본 생성된 시트 이름 변경 또는 처리
                ws_default = wb['Sheet']
                ws_default.title = name                

        if name in wb.sheetnames:           # 2. 시트가 이미 있으면 삭제
            del wb[name]                    # 또는 wb.remove(wb[sheet_name])
        
        new_ws = wb.create_sheet(name)      # 새로운 시트 생성
        base_cell = new_ws['A3']            # 1열 3행 
        set_title(new_ws, base_cell, code, name, inv_days, vo_rate)  

        total_sum.append(cal_roi(inv_days, new_ws, rows, wb))           
            
        wb.save("시뮬레이션결과.xlsx")          # 5. 저장


    for i in range(0, len(total_sum)):
        abc += total_sum[i][0]    # 평균 매수가
        # arr += total_sum[i][1]    # 평균 수익률
        roi += total_sum[i][2]    # 수익 금액

    msg = f'\nTotal ABC: {abc:,}원  ARR: {round((roi/abc)*100,2)}%  ROI: {roi:,}원'
    print(msg)
    tel_send(msg + '\n시뮬레이션결과.xlsx 업데이트 완료')
    return



import asyncio
# 실행 구간

async def main():
    await vobres_sim_excel("주식종목및코드 260510.xlsx", "코스피100")

if __name__ == '__main__':  
      asyncio.run(main())