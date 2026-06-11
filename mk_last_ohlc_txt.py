import json
import openpyxl
import os
import pandas as pd
import csv
# from mk_daypole_txt import read_excelfile
from util import read_datafile, read_excelfile

from tel_send import tel_send


def mk_last_ohlc_excel_to_txt(filename, option, stk_name_file):
    target_prc = {}

    script_dir = os.path.dirname(os.path.abspath(__file__))
    f_path = os.path.join(script_dir, filename)

    # script_dir = os.path.dirname(__file__)  # 현재 스크립트 파일이 있는 폴더
    # sim_filename = script_dir + "\\시뮬레이션결과.xlsx"
    
    wb = openpyxl.load_workbook(filename=f_path)

    sheet_name = option
    f_path = os.path.join(script_dir, stk_name_file)
    # excel_filename = stk_name_file
    df = read_excelfile(f_path, sheet_name)
    names = df['종목']      # 종목열만 추출 

    for i in range(len(names)):
        code_n_name = names[i]
        code = code_n_name[-6:]     # 종목코드 추출
        name = code_n_name[:-6]     # 종목명 추출
        print(f'{i+1} 종목코드: {code} 종목명: {name}')

        t_prc = {}
    
        ws = wb[name]

        if ws['A1'].value == '종목코드':        
            t_prc['종목코드'] = ws['B1'].value
            t_prc['종목명'] = ws['D1'].value
            t_prc['날짜'] = ws['B4'].value
            t_prc['시가'] = ws['C4'].value
            t_prc['고가'] = ws['D4'].value
            t_prc['저가'] = ws['E4'].value
            t_prc['현재가'] = ws['F4'].value
            t_prc['5MA'] = ws['G4'].value
            t_prc['10MA'] = ws['H4'].value
      
        if t_prc == None:
            msg = 'mk_last_ohlc.txt: Excel File Format Error'
            print(msg)
            tel_send(msg)
        else:    
            print(f'{i+1} ', t_prc)
            # 'stk'가 없으면 새 리스트를 만들고 추가
            target_prc.setdefault('stk', []).append(t_prc)
    # for
    print('\n', target_prc)

    try:
        fp = open('last_ohlc.txt', 'w', encoding='utf-8')
        fp.write(json.dumps(target_prc, indent=4, ensure_ascii=False))
        fp.close
    except Exception as e:
        print('last_ohlc.txt', e)



def mk_last_ohlc_csv_to_txt(stk_name_file, option):
    target_prc = {}

    script_dir = os.path.dirname(os.path.abspath(__file__))
    # script_dir = os.path.join(script_dir, 'daypole_data')

    sheet_name = option
    f_path = os.path.join(script_dir, stk_name_file)
    df = read_excelfile(f_path, sheet_name)
    names = df['종목']      # 종목열만 추출 

  
    script_dir = os.path.join(script_dir, 'daypole_data')
    for i in range(len(names)):
        code_n_name = names[i]
        code = code_n_name[-6:]     # 종목코드 추출
        name = code_n_name[:-6]     # 종목명 추출
        print(f'{i+1} 종목코드: {code} 종목명: {name}')

        f_path = os.path.join(script_dir, name + '.csv')
        with open(f_path, 'r', encoding='utf-8', newline='') as fp:
        # with open(f_path, 'r', newline='') as fp:
            reader = csv.reader(fp)
    
            csv_row = list(reader)
            
            t_prc = {}

            if csv_row[0][0] == '종목코드':       
                t_prc['종목코드'] = csv_row[0][1]
                t_prc['종목명'] = csv_row[0][3]
            
                t_prc['날짜'] = csv_row[3][1]   # 날짜
                t_prc['시가'] = csv_row[3][2]   # 시가
                t_prc['고가'] = csv_row[3][3]   # 고가
                t_prc['저가'] = csv_row[3][4]   # 저가
                t_prc['현재가'] = csv_row[3][5] # 현재가(종가)
                t_prc['5MA'] = csv_row[3][6]   # 5MA
                t_prc['10MA'] = csv_row[3][7]  # 10MA
        
                target_prc.setdefault('stk', []).append(t_prc)
                # print('\n', target_prc)
            if len(t_prc) == None:
                msg = 'mk_last_ohlc.txt: CSV File Format Error'
                print(msg)
                tel_send(msg)
            # with
    # for
    # print('\n', target_prc)

    try:
        fp = open('last_ohlc.txt', 'w', encoding='utf-8')
        fp.write(json.dumps(target_prc, indent=4, ensure_ascii=False))
        fp.close
    except Exception as e:
        print('last_ohlc.txt', e)



if __name__ == '__main__':
    # mk_last_ohlc_excel_to_txt('시뮬레이션결과.xlsx', '코스피100', '주식종목및코드 260510.xlsx')    
    mk_last_ohlc_csv_to_txt('주식종목및코드 260510.xlsx', '코스피100')    